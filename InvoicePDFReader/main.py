import os
import re
import openpyxl
from openpyxl.styles import Alignment
from PyPDF2 import PdfReader
import pdfplumber

def extract_text_from_pdf(pdf_path):
    # Use pdfplumber for more robust text extraction
    with pdfplumber.open(pdf_path) as pdf:
        text = ""
        for page in pdf.pages:
            text += page.extract_text() or ""
    return text

def clean_extracted_text(text):
    # More aggressive text cleaning
    # Remove newlines, extra spaces, and normalize
    cleaned_text = re.sub(r'\s+', ' ', text).strip()
    return cleaned_text

def parse_invoice_details(text):
    details = {}

    # Improved regex patterns for extraction
    patterns = {
        'Invoice Number': r'Invoice\s*number\s*([A-Za-z0-9-]+)',
        'Date of Issue': r'Date\s*of\s*issue\s*([A-Za-z]+\s+\d{1,2},\s+\d{4})',
        'Date Due': r'Date\s*due\s*([A-Za-z]+\s+\d{1,2},\s+\d{4})',
        'Bill To': r'Bill\s*to\s*(.*?)(?=\$|Subtotal|$)',
        'Subtotal': r'Subtotal\s*\$?\s*([0-9.]+)',
        'Tax': r'Tax\s*.*?\$\s*([0-9.]+)',
        'Total': r'Total\s*\$?\s*([0-9.]+)',
        'Amount Due': r'Amount\s*due\s*\$?\s*([0-9.]+)'
    }

    # Extract details with more flexible matching
    for key, pattern in patterns.items():
        match = re.search(pattern, text, re.IGNORECASE | re.DOTALL)
        if match:
            if key in ['Subtotal', 'Tax', 'Total', 'Amount Due']:
                details[key] = float(match.group(1))
            else:
                details[key] = match.group(1).strip()
        else:
            details[key] = "N/A"

    return details

def write_to_excel(details_list, output_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Invoice Details"

    headers = ['Invoice Number', 'Date of Issue', 'Date Due', 'Bill To', 'Subtotal', 'Tax', 'Total', 'Amount Due']
    for col_num, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.font = openpyxl.styles.Font(bold=True)

    for row_num, details in enumerate(details_list, start=2):
        for col_num, key in enumerate(headers, start=1):
            value = details.get(key, "N/A")
            ws.cell(row=row_num, column=col_num, value=value)

    wb.save(output_path)

def process_pdfs_in_directory(directory, output_excel_path):
    details_list = []
    for file_name in os.listdir(directory):
        if file_name.lower().endswith('.pdf'):
            pdf_path = os.path.join(directory, file_name)
            print(f"Processing {pdf_path}...")
            
            try:
                # Extract raw text
                raw_text = extract_text_from_pdf(pdf_path)
                cleaned_text = clean_extracted_text(raw_text)
                
                # Parse details
                details = parse_invoice_details(cleaned_text)
                print(f"Parsed details from {file_name}: {details}\n{'-'*50}")
                
                details_list.append(details)
            except Exception as e:
                print(f"Error processing {file_name}: {e}")
    
    write_to_excel(details_list, output_excel_path)

# Ensure required libraries are installed
# pip install openpyxl PyPDF2 pdfplumber

# Define paths
repository_path = '/Users/joshuabrooks/PythonProjectsPrivate-6/InvoicePDFReader'
output_excel_path = os.path.join(repository_path, 'Invoice_Details.xlsx')

# Process all PDF files in the directory
process_pdfs_in_directory(repository_path, output_excel_path)

print(f"Invoice details have been written to {output_excel_path}")